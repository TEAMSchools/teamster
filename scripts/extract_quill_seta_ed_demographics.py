# /// script
# requires-python = ">=3.13"
# dependencies = ["openpyxl>=3.1", "google-cloud-bigquery>=3.0"]
# ///
"""Build a de-identified demographics workbook for the Quill pilot roster.

The roster export keys students by Quill's own platform user ID, not by any
KIPP identifier, so the delivered workbook carries no KIPP identifier at all.
The mapping back to real students is written to a separate local key file that
is never committed and never transmitted.

See docs/reference/quill-seta-ed-demographics-extract.md.
"""

from __future__ import annotations

import argparse
import datetime
import json
import os
import re
import sys
from collections import Counter
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path

OUTPUT_COLUMNS: tuple[str, ...] = (
    "quill_student_id",
    "classroom_code",
    "teacher_code",
    "race_ethnicity",
    "gender",
    "meal_status",
    "iep_status",
    "mll_status",
    "demographics_source",
)

# Race categories follow rpt_gsheets__csgf_enrollment, but this script's
# hyphenated punctuation (BL-AA/AI-AN/NH-OPI) differs from csgf's slashes, so
# the strings won't join across the two outputs. GENDER_LABELS below instead
# follows int_extracts__student_enrollments.aligned_gender (csgf has no
# student gender map); MEAL_LABELS' Free/Reduced/Paid trichotomy is
# introduced by this script and has no prior model.
RACE_LABELS: dict[str, str] = {
    "A": "Asian",
    "B": "BL-AA",
    "H": "Hispanic or Latino",
    "I": "AI-AN",
    "M": "DTS",
    "N": "DTS",
    "P": "NH-OPI",
    "T": "2+ races",
    "W": "White",
    "Y": "DTS",
}
UNKNOWN_RACE = "DTS"

GENDER_LABELS: dict[str, str] = {"F": "Female", "M": "Male", "X": "Non-Binary"}

MEAL_LABELS: dict[str, str] = {
    "F": "Free",
    "FDC": "Free",
    "R": "Reduced",
    "P": "Paid",
}

IEP_LABELS: frozenset[str] = frozenset({"Has IEP", "No IEP"})
MLL_LABELS: frozenset[str] = frozenset({"ML", "Not ML"})

NOT_MATCHED = "not matched"
_SOURCE_PATTERN = re.compile(r"^SY\d{2}-\d{2}( \(fallback\))?$")
_CLASSROOM_CODE_PATTERN = re.compile(r"^C\d{2}$")
_TEACHER_CODE_PATTERN = re.compile(r"^T\d+$")


@dataclass(frozen=True)
class RosterRow:
    """One student-classroom pair as exported from the Quill platform."""

    quill_student_id: int
    student_email: str
    classroom_name: str
    teacher_name: str


@dataclass(frozen=True)
class Demographics:
    """One student's warehouse record for a single academic year."""

    academic_year: int
    student_number: int
    race_ethnicity: str
    gender: str
    meal_status: str
    iep_status: str
    mll_status: str


def school_year_label(academic_year: int) -> str:
    """2025 -> 'SY25-26'."""
    return f"SY{academic_year % 100:02d}-{(academic_year + 1) % 100:02d}"


def label_race(code: str | None) -> str:
    if code is None or code == "":
        return ""
    return RACE_LABELS.get(code.strip().upper(), UNKNOWN_RACE)


def label_gender(code: str | None) -> str:
    if code is None or code == "":
        return ""
    return GENDER_LABELS.get(code.strip().upper(), "")


def label_meal(code: str | None) -> str:
    if code is None or code == "":
        return ""
    return MEAL_LABELS.get(code.strip().upper(), "")


def assign_codes(
    names: Iterable[str],
    existing: Mapping[str, str],
    prefix: str,
    width: int,
) -> dict[str, str]:
    """Assign codes in order of first appearance, reusing prior assignments.

    Order of first appearance rather than alphabetical order: this script is
    public, and an alphabetical rule would let anyone who knows the real
    classroom or teacher names derive the mapping.
    """
    codes = dict(existing)
    used = [
        int(code[len(prefix) :])
        for code in codes.values()
        if code[len(prefix) :].isdigit()
    ]
    next_index = max(used, default=0) + 1

    for name in names:
        if name in codes:
            continue

        codes[name] = f"{prefix}{next_index:0{width}d}"
        next_index += 1

    return codes


def build_output_rows(
    roster: Sequence[RosterRow],
    demographics: Mapping[str, Demographics],
    pilot_year: int,
    classroom_codes: Mapping[str, str],
    teacher_codes: Mapping[str, str],
) -> list[dict[str, object]]:
    """One output row per roster row, carrying no identifier but the Quill ID."""
    rows: list[dict[str, object]] = []

    for entry in roster:
        record = demographics.get(entry.student_email)

        if record is None:
            source = NOT_MATCHED
        elif record.academic_year == pilot_year:
            source = school_year_label(record.academic_year)
        else:
            source = f"{school_year_label(record.academic_year)} (fallback)"

        rows.append(
            {
                "quill_student_id": entry.quill_student_id,
                "classroom_code": classroom_codes.get(entry.classroom_name, ""),
                "teacher_code": teacher_codes.get(entry.teacher_name, ""),
                "race_ethnicity": label_race(record.race_ethnicity if record else None),
                "gender": label_gender(record.gender if record else None),
                "meal_status": label_meal(record.meal_status if record else None),
                "iep_status": record.iep_status if record else "",
                "mll_status": record.mll_status if record else "",
                "demographics_source": source,
            }
        )

    return rows


def build_key_records(
    roster: Sequence[RosterRow],
    demographics: Mapping[str, Demographics],
    classroom_codes: Mapping[str, str],
    teacher_codes: Mapping[str, str],
) -> list[dict[str, object]]:
    """The re-identification key the agreement requires the supplier to retain."""
    records: list[dict[str, object]] = []

    for entry in roster:
        record = demographics.get(entry.student_email)

        records.append(
            {
                "quill_student_id": entry.quill_student_id,
                "student_email": entry.student_email,
                "student_number": record.student_number if record else None,
                "classroom_name": entry.classroom_name,
                "teacher_name": entry.teacher_name,
                "classroom_code": classroom_codes.get(entry.classroom_name, ""),
                "teacher_code": teacher_codes.get(entry.teacher_name, ""),
            }
        )

    return records


def validate_rows(
    rows: Sequence[Mapping[str, object]], roster: Sequence[RosterRow]
) -> None:
    """Whitelist every delivered cell. Raises ValueError on the first problem.

    Whitelisting rather than scanning for known names: a blacklist false-fires
    on surnames that collide with labels ('Free', 'Male') and silently misses
    anything not on the list.
    """
    if len(rows) != len(roster):
        raise ValueError(f"row count {len(rows)} does not match roster {len(roster)}")

    allowed_ids = {entry.quill_student_id for entry in roster}
    row_ids = {row["quill_student_id"] for row in rows}
    if len(row_ids) != len(allowed_ids):
        raise ValueError(
            f"distinct quill_student_id count {len(row_ids)} does not match "
            f"roster distinct count {len(allowed_ids)}"
        )

    allowed: dict[str, frozenset[str]] = {
        "race_ethnicity": frozenset(RACE_LABELS.values()) | {""},
        "gender": frozenset(GENDER_LABELS.values()) | {""},
        "meal_status": frozenset(MEAL_LABELS.values()) | {""},
        "iep_status": IEP_LABELS | {""},
        "mll_status": MLL_LABELS | {""},
    }

    for index, row in enumerate(rows):
        if tuple(row) != OUTPUT_COLUMNS:
            raise ValueError(f"row {index} columns are {tuple(row)}")

        if row["quill_student_id"] not in allowed_ids:
            raise ValueError(f"row {index} quill_student_id is not on the roster")

        classroom_code = str(row["classroom_code"])
        if not _CLASSROOM_CODE_PATTERN.match(classroom_code):
            raise ValueError(
                f"row {index} classroom_code {classroom_code!r} is not a code"
            )

        teacher_code = str(row["teacher_code"])
        if not _TEACHER_CODE_PATTERN.match(teacher_code):
            raise ValueError(f"row {index} teacher_code {teacher_code!r} is not a code")

        for column, vocabulary in allowed.items():
            value = str(row[column])
            if value not in vocabulary:
                raise ValueError(f"row {index} {column} disallowed value {value!r}")

        source = str(row["demographics_source"])
        if source != NOT_MATCHED and not _SOURCE_PATTERN.match(source):
            raise ValueError(f"row {index} demographics_source {source!r}")

        if source != NOT_MATCHED:
            for column in allowed:
                if str(row[column]) == "":
                    raise ValueError(
                        f"row {index} {column} is blank but demographics_source "
                        f"{source!r} indicates a match"
                    )


ROSTER_HEADERS: tuple[str, ...] = (
    "Studetn ID",  # upstream typo; matched verbatim so a fixed export fails loudly
    "Student Name",
    "Student Email",
    "Classroom Name",
    "Teacher Name",
    "Teacher Email",
)


def read_roster(path: Path) -> list[RosterRow]:
    """Read the Quill roster export, keeping only the columns we carry forward."""
    # trunk-ignore(pyright/reportMissingModuleSource): openpyxl is a PEP 723 script dep
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)

    header = tuple(str(value) if value is not None else "" for value in next(rows))
    if header[: len(ROSTER_HEADERS)] != ROSTER_HEADERS:
        raise ValueError(f"unexpected roster header {header!r}")

    roster: list[RosterRow] = []
    for row in rows:
        raw_id = row[0]
        if raw_id is None:
            continue

        # openpyxl types a cell as a wide union (formula, datetime, rich text).
        # Narrow before int() so an unexpected ID cell fails loudly here rather
        # than raising something obscure, and so pyright can check the call.
        # bool is an int subclass, so True would otherwise coerce to id 1; a
        # fractional float would truncate. Both are rejected, not coerced.
        if isinstance(raw_id, bool) or not isinstance(raw_id, int | float | str):
            raise ValueError(f"unexpected roster id cell {raw_id!r}")

        if isinstance(raw_id, float) and not raw_id.is_integer():
            raise ValueError(f"fractional roster id cell {raw_id!r}")

        # str(None) is the string "None", so a blank email would become the
        # literal "none" and silently miss every warehouse match. Fail instead.
        if row[2] is None or row[3] is None or row[4] is None:
            raise ValueError(f"roster row for id {raw_id!r} has a blank required cell")

        roster.append(
            RosterRow(
                quill_student_id=int(raw_id),
                student_email=str(row[2]).strip().lower(),
                classroom_name=str(row[3]).strip(),
                teacher_name=str(row[4]).strip(),
            )
        )

    workbook.close()

    return roster


def read_codebook(path: Path) -> tuple[dict[str, str], dict[str, str]]:
    """Prior code assignments, so a rerun keeps the same codes."""
    if not path.exists():
        return {}, {}

    stored = json.loads(path.read_text(encoding="utf-8"))

    return stored.get("classroom_codes", {}), stored.get("teacher_codes", {})


def write_key_file(
    path: Path,
    classroom_codes: Mapping[str, str],
    teacher_codes: Mapping[str, str],
    key_records: Sequence[Mapping[str, object]],
) -> None:
    """Write the retained re-identification key. Local only, never transmitted.

    Merges this run's student records into any prior key file by the roster
    row's natural key, (quill_student_id, classroom_name), rather than
    overwriting the students array outright or merging on quill_student_id
    alone. quill_student_id alone would collapse a student who appears in two
    Quill sections down to a single record, silently dropping one of their
    two classroom associations. Overwriting outright would drop any student
    who left the pilot roster between the pre- and post-period deliveries,
    making the already-delivered file's rows non-re-identifiable from the one
    artifact the agreement obliges the supplier to retain. The current run's
    record wins on conflict; prior-only records are preserved -- including a
    prior record for a since-renamed section, which is deliberate: a rename
    mints a new (id, name) pair rather than overwriting the old one, so the
    old code stays resolvable from the retained key. A `.bak` snapshot of the
    prior file is written alongside it before it is replaced, and the new
    content is written via a temp file plus an atomic replace so a crash
    mid-write cannot corrupt the only copy.
    """
    merged_students: dict[tuple[object, object], dict[str, object]] = {}

    if path.exists():
        prior = json.loads(path.read_text(encoding="utf-8"))
        for record in prior.get("students", []):
            key = (record["quill_student_id"], record["classroom_name"])
            merged_students[key] = dict(record)

        path.with_name(f"{path.name}.bak").write_bytes(path.read_bytes())

    for record in key_records:
        key = (record["quill_student_id"], record["classroom_name"])
        merged_students[key] = dict(record)

    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_name(f"{path.name}.tmp")
    temp_path.write_text(
        json.dumps(
            {
                "classroom_codes": dict(classroom_codes),
                "teacher_codes": dict(teacher_codes),
                "students": [
                    merged_students[key] for key in sorted(merged_students, key=str)
                ],
            },
            indent=2,
            sort_keys=True,
        ),
        encoding="utf-8",
    )
    os.replace(temp_path, path)


DATA_DICTIONARY: tuple[tuple[str, str, str, str], ...] = (
    (
        "quill_student_id",
        "Quill platform user ID for the student, unchanged from the roster export.",
        "Quill roster export, column 'Studetn ID'",
        "Integer. Join key to Quill platform data. Not a school district identifier.",
    ),
    (
        "classroom_code",
        "Pseudonymous code for the Quill classroom section.",
        "Assigned by this script",
        "C01 through C09. Mapping to section names is retained by the data supplier.",
    ),
    (
        "teacher_code",
        "Pseudonymous code for the teacher of record for the section.",
        "Assigned by this script",
        "T1, T2. Maps 1:1 to school. Mapping to names is retained by the data supplier.",
    ),
    (
        "race_ethnicity",
        "Race and ethnicity as recorded in the student information system.",
        "int_extracts__student_enrollments.race_ethnicity",
        "Asian, BL-AA, Hispanic or Latino, AI-AN, NH-OPI, 2+ races, White, DTS. "
        "DTS means declined to state. Blank when the student was not matched.",
    ),
    (
        "gender",
        "Gender as recorded in the student information system.",
        "int_extracts__student_enrollments.gender",
        "Female, Male, Non-Binary. Blank when the student was not matched.",
    ),
    (
        "meal_status",
        "National School Lunch Program eligibility for the school year.",
        "int_extracts__student_enrollments.lunch_status",
        "Free, Reduced, Paid. Direct certification is reported as Free. "
        "Blank when the student was not matched.",
    ),
    (
        "iep_status",
        "Whether the student had an active individualized education program.",
        "int_extracts__student_enrollments.iep_status",
        "Has IEP, No IEP. Blank when the student was not matched.",
    ),
    (
        "mll_status",
        "Multilingual learner status.",
        "int_extracts__student_enrollments.ml_status",
        "ML, Not ML. Blank when the student was not matched.",
    ),
    (
        "demographics_source",
        "Which school year's enrollment record supplied this row's values.",
        "Assigned by this script",
        "SY25-26 for the pilot year. 'SYxx-yy (fallback)' when the student was "
        "not enrolled in the pilot year and an earlier year was used. "
        "'not matched' when no enrollment record was found at all.",
    ),
)

_DICTIONARY_HEADERS = ("column", "definition", "source", "coding")


def write_workbook(path: Path, rows: Sequence[Mapping[str, object]]) -> None:
    """Write the two-sheet deliverable."""
    # trunk-ignore(pyright/reportMissingModuleSource): openpyxl is a PEP 723 script dep
    import openpyxl

    workbook = openpyxl.Workbook()

    # worksheets[0] rather than .active: pyright types .active as Optional and
    # trunk runs pyright over scripts/.
    data_sheet = workbook.worksheets[0]
    data_sheet.title = "student_demographics"
    data_sheet.append(list(OUTPUT_COLUMNS))
    for row in rows:
        data_sheet.append([row[column] for column in OUTPUT_COLUMNS])

    dictionary_sheet = workbook.create_sheet("data_dictionary")
    dictionary_sheet.append(list(_DICTIONARY_HEADERS))
    for entry in DATA_DICTIONARY:
        dictionary_sheet.append(list(entry))

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def read_workbook_rows(path: Path) -> list[dict[str, object]]:
    """Read the delivered sheet back, so validation runs against the real file.

    An in-memory check cannot catch a serialization mistake, and the file is
    what gets sent. openpyxl reads a blank cell as None; coerce to '' so the
    whitelist in validate_rows sees the same vocabulary it wrote.
    """
    # trunk-ignore(pyright/reportMissingModuleSource): openpyxl is a PEP 723 script dep
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["student_demographics"]
    rows = sheet.iter_rows(values_only=True)

    header = tuple(next(rows))
    if header != OUTPUT_COLUMNS:
        raise ValueError(f"unexpected workbook header {header!r}")

    parsed: list[dict[str, object]] = [
        {
            column: "" if value is None else value
            for column, value in zip(OUTPUT_COLUMNS, row, strict=True)
        }
        for row in rows
    ]

    workbook.close()

    return parsed


SOURCE_TABLE = "teamster-332318.kipptaf_extracts.int_extracts__student_enrollments"


def build_query(table: str) -> str:
    """One row per roster email: the pilot year if present, else the most recent.

    rn_year = 1 picks the student's primary enrollment stint within a year, so
    a mid-year school move does not produce two rows for one student-year.
    """
    # trunk-ignore(bandit/B608): table is a module constant, not user input; emails/pilot_year are bound as query parameters
    return f"""
        with
            ranked as (
                select
                    lower(student_email) as student_email,
                    student_number,
                    academic_year,
                    race_ethnicity,
                    gender,
                    lunch_status,
                    iep_status,
                    ml_status,
                    row_number() over (
                        partition by lower(student_email)
                        order by academic_year desc, student_number desc
                    ) as year_rank,
                from `{table}`
                where
                    lower(student_email) in unnest(@emails)
                    and academic_year <= @pilot_year
                    and rn_year = 1
            )
        select * except (year_rank)
        from ranked
        where year_rank = 1
    """


def fetch_demographics(
    client,
    emails: Sequence[str],
    pilot_year: int,
    table: str = SOURCE_TABLE,
) -> dict[str, Demographics]:
    """Query the warehouse for one demographics record per roster email."""
    from google.cloud import bigquery

    job_config = bigquery.QueryJobConfig(
        query_parameters=[
            bigquery.ArrayQueryParameter("emails", "STRING", list(emails)),
            bigquery.ScalarQueryParameter("pilot_year", "INT64", pilot_year),
        ]
    )

    results: dict[str, Demographics] = {}
    for row in client.query(build_query(table), job_config=job_config).result():
        results[row["student_email"]] = Demographics(
            academic_year=row["academic_year"],
            student_number=row["student_number"],
            race_ethnicity=row["race_ethnicity"],
            gender=row["gender"],
            meal_status=row["lunch_status"],
            iep_status=row["iep_status"],
            mll_status=row["ml_status"],
        )

    return results


def _report_cell_counts(rows: Sequence[Mapping[str, object]]) -> None:
    """Print the counts that decide whether this file is safe to send."""
    for column in (
        "race_ethnicity",
        "gender",
        "meal_status",
        "iep_status",
        "mll_status",
        "demographics_source",
    ):
        counts = Counter(str(row[column]) for row in rows)
        print(f"\n{column}:")
        for value, count in sorted(counts.items(), key=lambda item: -item[1]):
            flag = "  <-- singleton" if count == 1 else ""
            print(f"  {value or '(blank)':<24} {count:>4}{flag}")


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--roster", type=Path, required=True, help="roster .xlsx path")
    parser.add_argument(
        "--output-dir", type=Path, required=True, help="directory for the workbook"
    )
    parser.add_argument(
        "--key-file",
        type=Path,
        required=True,
        help="retained re-identification key (JSON); never commit or transmit",
    )
    parser.add_argument("--pilot-year", type=int, default=2025)
    parser.add_argument("--gcp-project", default="teamster-332318")
    args = parser.parse_args(argv)

    from google.cloud import bigquery

    roster = read_roster(args.roster)
    print(
        f"roster: {len(roster)} rows, {len({e.quill_student_id for e in roster})} students"
    )

    # validate_rows compares len(rows) to len(roster), so an empty roster
    # validates trivially and would ship an empty workbook as a success.
    if not roster:
        raise ValueError(f"roster at {args.roster} produced no rows")

    prior_classrooms, prior_teachers = read_codebook(args.key_file)
    classroom_codes = assign_codes(
        (entry.classroom_name for entry in roster), prior_classrooms, "C", 2
    )
    teacher_codes = assign_codes(
        (entry.teacher_name for entry in roster), prior_teachers, "T", 1
    )

    # A renamed section mints a fresh code with nothing else printed, which
    # silently breaks cross-delivery linkage -- report reuse so that shows up.
    classroom_names = {entry.classroom_name for entry in roster}
    teacher_names = {entry.teacher_name for entry in roster}
    reused_classrooms = len(classroom_names & prior_classrooms.keys())
    reused_teachers = len(teacher_names & prior_teachers.keys())
    print(
        f"classroom codes: {reused_classrooms} reused, "
        f"{len(classroom_names) - reused_classrooms} new; "
        f"teacher codes: {reused_teachers} reused, "
        f"{len(teacher_names) - reused_teachers} new"
    )

    emails = sorted({entry.student_email for entry in roster})
    demographics = fetch_demographics(
        bigquery.Client(project=args.gcp_project), emails, args.pilot_year
    )
    print(f"matched {len(demographics)} of {len(emails)} emails")

    # 'not matched' is a legal demographics_source, so a total fetch failure
    # would otherwise produce a well-formed workbook with no real data in it.
    if not demographics:
        raise ValueError(
            f"no warehouse match for any of {len(emails)} roster emails; "
            "refusing to write an all-unmatched deliverable"
        )

    rows = build_output_rows(
        roster, demographics, args.pilot_year, classroom_codes, teacher_codes
    )
    validate_rows(rows, roster)

    extract_date = datetime.date.today().isoformat()
    workbook_path = (
        args.output_dir / f"quill_seta_ed_student_demographics_{extract_date}.xlsx"
    )
    # openpyxl's reader rejects a path whose extension isn't .xlsx/.xlsm/etc
    # regardless of content, so the temp name keeps the real suffix rather
    # than appending one.
    temp_workbook_path = workbook_path.with_name(f"{workbook_path.stem}.tmp.xlsx")

    # Write to a temp path in the same directory and validate there first.
    # The workbook only lands on its final, sendable name once every check
    # below has passed -- otherwise a validation failure leaves a rejected
    # file sitting under the exact name an operator could attach to an email.
    try:
        write_workbook(temp_workbook_path, rows)

        # Re-validate from disk. The check above proves the rows were built
        # right; this one proves the file that gets sent is the file we
        # built. The equality check below catches a scrambled row order that
        # the whitelist alone would accept -- valid cells attached to the
        # wrong quill_student_id.
        reread_rows = read_workbook_rows(temp_workbook_path)
        validate_rows(reread_rows, roster)
        if reread_rows != rows:
            raise ValueError(
                "workbook rows do not match the rows that were built, even "
                "though each row individually passed validation"
            )
    except Exception:
        temp_workbook_path.unlink(missing_ok=True)
        raise

    # Write the retained key BEFORE the workbook reaches its sendable name.
    # write_key_file does much more than os.replace (parses prior JSON, writes
    # a .bak, merges records, writes its own temp file) and so has more ways
    # to fail. If it raised after the rename below, a correctly-named,
    # fully-formed workbook would sit on disk with no matching
    # re-identification key -- the artifact the agreement requires the
    # supplier to retain. The reverse failure is the one we accept: if
    # os.replace fails after this call, the key records a delivery that
    # never shipped, which is harmless because the key is cumulative and a
    # rerun rewrites it. Do not swap this order back.
    write_key_file(
        args.key_file,
        classroom_codes,
        teacher_codes,
        build_key_records(roster, demographics, classroom_codes, teacher_codes),
    )

    os.replace(temp_workbook_path, workbook_path)

    _report_cell_counts(rows)
    print(f"\nworkbook: {workbook_path}")
    print(f"key file (do not send): {args.key_file}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
