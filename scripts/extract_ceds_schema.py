# /// script
# requires-python = ">=3.13"
# dependencies = ["openpyxl>=3.1"]
# ///

"""Extract CEDS element names and definitions.

Fetches the CEDS-Elements XLSX workbook from GitHub, joins the Properties
and Classes sheets via `Domain Includes` IRI, and writes a vendored CSV
in the same shape as `edfi-v6.1.0-attributes.csv`.

Output: docs/superpowers/specs/ceds-v14.0.0.0-attributes.csv

Usage:
    uv run scripts/extract_ceds_schema.py

Design reference:
    docs/superpowers/specs/2026-04-15-column-naming-audit.md
"""

from __future__ import annotations

import csv
import re
import urllib.request
from io import BytesIO
from pathlib import Path
from typing import TextIO

# trunk-ignore(pyright/reportMissingModuleSource): openpyxl is a PEP 723 script dep
from openpyxl import load_workbook

CEDS_VERSION = "V14.0.0.0"
CEDS_XLSX_URL = (
    f"https://raw.githubusercontent.com/CEDStandards/CEDS-Elements/"
    f"master/CEDS-{CEDS_VERSION}.xlsx"
)

CSV_FIELDS: tuple[str, ...] = (
    "entity",
    "attribute_camel",
    "attribute_snake",
    "description",
)

OUTPUT_CSV = Path(f"docs/superpowers/specs/ceds-{CEDS_VERSION.lower()}-attributes.csv")

_PASCAL_RE = re.compile(r"(?<=[a-z0-9])(?=[A-Z])|(?<=[A-Z])(?=[A-Z][a-z])")


def _pascal_to_snake(name: str) -> str:
    """Convert PascalCase or camelCase to snake_case."""
    return _PASCAL_RE.sub("_", name).lower()


def _fetch_workbook() -> BytesIO:
    """Fetch the CEDS XLSX from GitHub as in-memory bytes."""
    # trunk-ignore(bandit/B310): URL is a hardcoded https GitHub raw path
    with urllib.request.urlopen(CEDS_XLSX_URL) as resp:
        return BytesIO(resp.read())


def _load_class_index(wb) -> dict[str, str]:
    """Build a lookup: Class IRI → Class Technical Name.

    Skips deprecated classes (detected via `(Deprecated)` prefix on the
    Technical Name, since the Classes sheet has no owl:deprecated column).
    """
    ws = wb["Classes"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    iri_idx = header.index("IRI")
    tn_idx = header.index("Technical Name (skos:notation)")

    index: dict[str, str] = {}
    for row in rows:
        iri = row[iri_idx]
        tn = row[tn_idx]
        if not iri or not tn:
            continue
        if str(tn).startswith("(Deprecated)"):
            continue
        index[str(iri)] = str(tn)
    return index


def _extract_attributes(wb, class_index: dict[str, str]) -> list[dict[str, str]]:
    """Extract (entity, attribute, description) triples from the Properties sheet.

    Skips deprecated properties. Properties whose Domain Includes IRI
    resolves to a non-indexed (deprecated) class are also skipped.
    """
    ws = wb["Properties"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    tn_idx = header.index("Technical Name (skos:notation)")
    def_idx = header.index("Definition (skos:definition)")
    di_idx = header.index("Domain Includes (schema:domainIncludes)")
    dep_idx = header.index("Deprecated (owl:deprecated)")

    out: list[dict[str, str]] = []
    for row in rows:
        if row[dep_idx]:
            continue
        tn = row[tn_idx]
        domain_iri = row[di_idx]
        if not tn or not domain_iri:
            continue
        entity = class_index.get(str(domain_iri))
        if not entity:
            continue
        camel = str(tn)
        out.append(
            {
                "entity": entity,
                "attribute_camel": camel,
                "attribute_snake": _pascal_to_snake(camel),
                "description": str(row[def_idx] or ""),
            }
        )
    return out


def _write_csv(rows: list[dict[str, str]], fh: TextIO) -> None:
    writer = csv.DictWriter(fh, fieldnames=list(CSV_FIELDS), lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow({k: row.get(k, "") for k in CSV_FIELDS})


def main() -> None:
    buf = _fetch_workbook()
    wb = load_workbook(buf, read_only=True, data_only=True)
    class_index = _load_class_index(wb)
    rows = _extract_attributes(wb, class_index)

    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_CSV.open("w", encoding="utf-8", newline="") as fh:
        _write_csv(rows, fh)

    print(f"Wrote {len(rows)} attributes to {OUTPUT_CSV}")


if __name__ == "__main__":
    main()
